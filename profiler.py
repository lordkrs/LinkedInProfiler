import os
import xmltodict
import datetime
import json
import uuid
import xlsxwriter
import openpyxl
from bottle import abort, request, static_file, run, route
import sys
import zipfile
from googlesearch import search

temp_path = os.path.dirname(os.path.abspath(__file__)) + os.path.sep + "tmp"
if not os.path.exists(temp_path):
    os.makedirs(temp_path)


SHEET_LIMIT = 50000
MAX_SHEETS_PER_XLS = 7
LINKEDIN_PROFILER_HEADER = ["First Name", "Last Name", "Company Name","Title","Search Field", "LinkedIn Link", "Other Links"]


def zipper(zip_file_name, files):
    zip_file_name = '{}{}{}.zip'.format(temp_path, os.path.sep, zip_file_name)
    print("Zipping {} xlsx files to {}".format(len(files), zip_file_name))
    with zipfile.ZipFile(zip_file_name,'w') as zip_:
        for file_ in files:
            zip_.write(temp_path + os.path.sep + file_)
    print("file:{}".format(zip_file_name))
    return os.path.basename(zip_file_name)
        

def create_xlsx(data=None, data_list=[], local=False,headers=LINKEDIN_PROFILER_HEADER,sheet_limit=SHEET_LIMIT):
    main_file_name = str(uuid.uuid4())+".xlsx"
    sheet_number = 1
    workbook = xlsxwriter.Workbook(temp_path + os.path.sep + main_file_name)
    worksheet = workbook.add_worksheet(name="Sheet{}".format(sheet_number))
    row = 0
    col = 0
    total_files = [main_file_name]
    for header in headers:
        worksheet.write(row, col, header)
        col += 1

    if not local:
        for col_data in data:
            row += 1
            col = 0
            for header in headers:
                #print("header--> {}:data-->{}:type--->{}".format(header,col_data[header],type(col_data[header])))
                worksheet.write(row, col, col_data[header])
                col += 1

            if sheet_number < MAX_SHEETS_PER_XLS:
                if row >= sheet_limit:
                    row = 0
                    col = 0
                    sheet_number += 1
                    worksheet = workbook.add_worksheet(name="Sheet{}".format(sheet_number))
                    for header in headers:
                        worksheet.write(row, col, header)
                        col += 1
            else:
                if row < sheet_limit:
                    continue
                workbook.close() 
                file_name = "{}_part-{}.xlsx".format(os.path.splitext(main_file_name)[0], len(total_files))
                sheet_number = 1
                workbook = xlsxwriter.Workbook(temp_path + os.path.sep + file_name)
                worksheet = workbook.add_worksheet(name="Sheet{}".format(sheet_number))
                row = 0
                col = 0
                for header in headers:
                    worksheet.write(row, col, header)
                    col += 1
                total_files.append(file_name)
                
    else:
        for data in data_list:
            for col_data in data:
                row += 1
                col = 0
                for header in headers:
                    #print("header--> {}:data-->{}:type--->{}".format(header,col_data[header],type(col_data[header])))
                    worksheet.write(row, col, col_data[header])
                    col += 1

                if sheet_number < MAX_SHEETS_PER_XLS:
                    if row >= sheet_limit:
                        row = 0
                        col = 0
                        sheet_number += 1
                        worksheet = workbook.add_worksheet(name="Sheet{}".format(sheet_number))
                        for header in headers:
                            worksheet.write(row, col, header)
                            col += 1
                else:
                    if row < sheet_limit:
                        continue
                    workbook.close() 
                    file_name = "{}_part-{}.xlsx".format(os.path.splitext(main_file_name)[0], len(total_files))
                    sheet_number = 1
                    workbook = xlsxwriter.Workbook(temp_path + os.path.sep + file_name)
                    worksheet = workbook.add_worksheet(name="Sheet{}".format(sheet_number))
                    row = 0
                    col = 0
                    for header in headers:
                        worksheet.write(row, col, header)
                        col += 1
                    total_files.append(file_name)  

    workbook.close()      

    if len(total_files) == 1:
        return main_file_name
    else:
        return zipper(os.path.splitext(main_file_name)[0], total_files)

def xml_to_json(xml):
    ''' This API converts xml data to json
        parameters:
        ---------------
        xml (str) : Complete xml data read from website or file

        return:
        ---------------
        json(dict) : returns valid dictionary
    '''
    json_string = json.dumps(xmltodict.parse(xml))
    json_data = json.loads(json_string)
    return json.dumps(json_data)

@route('/upload', method='POST')
def do_upload():

    upload     = request.files.get('upload')
    name, ext = os.path.splitext(upload.filename)
    
    if ext in ('.png','.jpg','.jpeg'):
        return 'File extension not allowed.'

    upload.save(temp_path) # appends upload.filename automatically
    
    sheet_len = int(request.forms.get("sheet_len")) if request.forms.get('sheet_len') else SHEET_LIMIT 
    
    try:
        xlsx_file_path = os.path.join(temp_path, upload.filename)
        xlsx_data = []
        wb_obj = openpyxl.load_workbook(xlsx_file_path)
        sheet_obj = wb_obj.active
        for i in range(1, sheet_obj.max_row+1):
            row_data = {}
            for j in range(1, sheet_obj.max_column+1):
                row_data[sheet_obj.cell(row = 1, column = j).value] = sheet_obj.cell(row = i+1, column = j).value
            xlsx_data.append(row_data)
        
        os.remove(xlsx_file_path)
        ids_return_data = {"ids_info":{},"count":0}

        xlsx_data_list = []
         
        for column_data in xlsx_data:
            print("\n\n{} of {}\n\n".format(xlsx_data.index(column_data)+1, len(xlsx_data)))
            first_name = column_data["First_Name"] if column_data.get("First_Name") else ""
            last_name = column_data["Last_Name"] if column_data.get("Last_Name") else ""
            company_name = column_data["Company_Name"] if column_data.get("Company_Name") else ""
            job_title = column_data["Job_Title"] if column_data.get("Job_Title") else ""


            search_field = "" + company_name
            if first_name == "Unknown" or last_name == "Unknown":
                continue
            else:
                search_field += " {} {}".format(first_name, last_name) 

            if job_title:
                search_field += " {}".format(job_title)
            
            search_field += " linkedIn"
             
            results = search(search_field)


            if len(results) != 0:

                url = results[0]
                for res in results:
                    if first_name.lower() in res.lower():
                        url = "" + res
                    elif last_name.lower() in res.lower():
                        url = "" + res
                        
                xlsx_data_list.append({"Search Field": search_field, "Title": job_title,"LinkedIn Link": url, "Company Name":company_name, "First Name":first_name, "Last Name": last_name, "Other Links": ", ".join(results)})
            

        if len(xlsx_data_list) != 0:
            file_path = create_xlsx(data=xlsx_data_list, local=False, headers=LINKEDIN_PROFILER_HEADER, sheet_limit=sheet_len)
            return static_file(file_path, temp_path, download=file_path)
    except Exception as ex:
        print("Exception in upload:{}".format(ex))
        abort(500, "Exception occurred: {}".format(ex))        

        
@route("/clear_tmp",method="POST")
def clear_tmp():
    files = os.listdir(temp_path)
    for file_ in files:
        os.remove(file_)    

@route("/css/<css_file>")
def serve_css(css_file):
    return static_file(css_file, os.path.dirname(os.path.abspath(__file__))+os.path.sep+"css")

@route("/")
def serve_web():
    return static_file("index.html", os.path.dirname(os.path.abspath(__file__)))

if __name__ == "__main__":
    run(host="0.0.0.0",port=8090)
